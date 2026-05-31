"""Generate the master accounts Excel file for all promotion channels."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Styles ──────────────────────────────────────────────
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(bold=True, size=11, color="2F5496")
warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border


def add_row(ws, row, data):
    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ═══════════════════════════════════════════════════════════
# SHEET 1: ALL ACCOUNTS
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Accounts Master"
cols = 8
headers = ["Platform", "Username/Handle", "Email", "Password", "Product", "Purpose", "API Key/Token", "Notes"]

r = 1
style_header(ws, r, cols)
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
r += 1

# ── Reddit ──
style_section(ws, r, cols, "REDDIT ACCOUNTS")
r += 1
add_row(ws, r, ["Reddit", "Khavel_dev", "ceja_pontejos@hotmail.com", "(see password manager)", "DevAI Semanal",
    "Warm-up + promo: r/ClaudeAI, r/microsaas, r/SideProject, r/SaaS", "",
    "Created 2026-05-26. Profile: seo/.reddit-profile/. 4 comments posted day 1."])
r += 1
add_row(ws, r, ["Reddit", "StatLineNerd", "StatLineNerd@gmail.com", "(see password manager)", "NbaPropLab",
    "Warm-up + promo: r/sportsbook, r/NBAbetting, r/fantasybball", "",
    "Created 2026-05-26. Needs Playwright profile setup."])
r += 1
add_row(ws, r, ["Reddit", "xGNerd", "cejatron121@gmail.com", "(see password manager)", "FutPicks",
    "Warm-up + promo: r/SoccerBetting, r/FootballTips", "",
    "Created 2026-05-26. Needs Playwright profile setup."])
r += 1

# ── Twitter/X ──
style_section(ws, r, cols, "TWITTER/X ACCOUNTS")
r += 1
add_row(ws, r, ["Twitter/X", "@DevAISemanal", "??? (fill in)", "??? (fill in)", "DevAI Semanal",
    "AI dev tips, newsletter teasers, build-in-public", "",
    "Created 2026-05-26. Verified."])
r += 1
add_row(ws, r, ["Twitter/X", "@StatLineNerd", "??? (fill in)", "??? (fill in)", "NbaPropLab",
    "NBA prop picks, spider charts, analysis threads", "",
    "Created 2026-05-26."])
r += 1
add_row(ws, r, ["Twitter/X", "@FutProbLab", "??? (fill in)", "??? (fill in)", "FutPicks",
    "Football picks, weekend previews, bilingual es/en", "",
    "Created 2026-05-26. Verified."])
r += 1

# ── Dev.to ──
style_section(ws, r, cols, "DEV.TO")
r += 1
add_row(ws, r, ["Dev.to", "Khavel", "ceja_pontejos@hotmail.com", "(GitHub OAuth)", "DevAI Semanal",
    "Cross-post newsletter editions as blog articles",
    "(API key — see password manager)",
    "Account since Aug 2022. 4 badges. Published 1 article (May 25)."])
r += 1

# ── Hashnode ──
style_section(ws, r, cols, "HASHNODE")
r += 1
add_row(ws, r, ["Hashnode", "Khavel (via GitHub OAuth)", "ceja_pontejos@hotmail.com", "(GitHub OAuth)", "DevAI Semanal",
    "Cross-post newsletter editions as blog articles",
    "(PAT — see password manager)",
    "Created 2026-05-26. Logged in via GitHub. PAT above."])
r += 1

# ── Telegram ──
style_section(ws, r, cols, "TELEGRAM BOTS")
r += 1
add_row(ws, r, ["Telegram", "NbaPropLab bot", "", "", "NbaPropLab",
    "Auto-publish picks + Telegraph pages via NBAv3.Telegram", "",
    "Bot configured in NBAv3.Telegram service. Check .env for bot token."])
r += 1
add_row(ws, r, ["Telegram", "FootballLab bot", "", "", "FutPicks",
    "Auto-publish picks via FootballLab.Telegram", "",
    "Bot configured in FootballLab.Telegram. Check .env for bot token."])
r += 1

# ── GitHub ──
style_section(ws, r, cols, "GITHUB")
r += 1
add_row(ws, r, ["GitHub", "Khavel", "ceja_pontejos@hotmail.com", "", "All",
    "Source repos, awesome-list PRs, open source", "",
    "Main GitHub account. PRs submitted to 5 awesome-lists."])
r += 1

# ── Directory submissions ──
style_section(ws, r, cols, "NEWSLETTER DIRECTORY SUBMISSIONS (DevAI Semanal)")
r += 1
dirs = [
    ["Paved", "Submitted 2026-05-26. DA ~55, dofollow."],
    ["Crunchbase", "Submitted 2026-05-26. DA 91, dofollow. Pending editorial review."],
    ["InboxReads", "Submitted 2026-05-25. No account created."],
    ["LetterList", "Submitted 2026-05-25 via Airtable. Pending review."],
    ["Rad Letters", "Submitted 2026-05-25 via React form. Pending review."],
    ["StackLetter", "Submitted 2026-05-25. Pending review."],
    ["Product Hunt", "Launch checked 2026-05-26."],
    ["G2", "BLOCKED: needs business Google account. DA 91."],
    ["AlternativeTo", "WAITING: eligible after June 1. DA 70+."],
    ["Indie Hackers", "NOT DONE: needs manual logo upload. DA 65+."],
    ["Find Your Newsletter", "BLOCKED: reCAPTCHA issue."],
    ["PostApex", "Status unknown. Check."],
]
for d in dirs:
    add_row(ws, r, [d[0], "DevAI Semanal", "ceja_pontejos@hotmail.com", "", "DevAI Semanal",
        "Newsletter directory listing", "", d[1]])
    r += 1

# ── GitHub PRs ──
style_section(ws, r, cols, "GITHUB AWESOME-LIST PRs")
r += 1
prs = [
    ["awesome-newsletters", "PR #332 to zudochkin/awesome-newsletters"],
    ["awesome-ai-newsletters", "PR #39"],
    ["awesome-weekly", "PR #56"],
    ["developer-newsletters", "PR #31"],
    ["awesome-web-newsletters", "PR #8"],
]
for pr in prs:
    add_row(ws, r, ["GitHub PR", f"Khavel/{pr[0]}", "", "", "DevAI Semanal",
        f"Awesome-list backlink: {pr[1]}", "", "Submitted. Pending merge."])
    r += 1

# Column widths
widths = [16, 28, 30, 14, 16, 52, 40, 58]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════
# SHEET 2: PENDING SETUP ACTIONS
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pending Setup")
cols2 = 5
headers2 = ["#", "Action", "Product", "Priority", "Status"]
style_header(ws2, 1, cols2)
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)

pending = [
    [1, "Create Reddit account for FutPicks (xGNerd)", "FutPicks", "HIGH", "DONE"],
    [2, "Create Twitter/X account for FutPicks (@FutProbLab)", "FutPicks", "HIGH", "DONE"],
    [3, "Fill in Twitter emails/passwords in Accounts Master sheet", "All", "HIGH", "NOT DONE"],
    [4, "Fill in Hashnode username in Accounts Master sheet", "DevAI", "MEDIUM", "NOT DONE"],
    [5, "Log into StatLineNerd on Reddit via Playwright (setup profile)", "NbaPropLab", "HIGH", "NOT DONE"],
    [6, "Log into @DevAISemanal on Twitter via Playwright (setup profile)", "DevAI", "HIGH", "NOT DONE"],
    [7, "Log into @StatLineNerd on Twitter via Playwright (setup profile)", "NbaPropLab", "HIGH", "NOT DONE"],
    [8, "Build Twitter automation script (Playwright approach)", "All", "HIGH", "NOT BUILT"],
    [9, "Build Reddit automation for StatLineNerd (clone Khavel_dev setup)", "NbaPropLab", "HIGH", "NOT BUILT"],
    [10, "Build Dev.to cross-post script (API)", "DevAI", "MEDIUM", "NOT BUILT"],
    [11, "Build Hashnode cross-post script (GraphQL API)", "DevAI", "MEDIUM", "NOT BUILT"],
    [12, "Submit to G2 (needs business Google account)", "DevAI", "MEDIUM", "BLOCKED"],
    [13, "Submit to AlternativeTo (after June 1)", "DevAI", "LOW", "WAITING"],
    [14, "Submit to Indie Hackers (manual logo upload)", "DevAI", "MEDIUM", "NOT DONE"],
    [15, "Complete Find Your Newsletter (reCAPTCHA)", "DevAI", "LOW", "BLOCKED"],
    [16, "Check PostApex status", "DevAI", "LOW", "NOT DONE"],
]
for i, row_data in enumerate(pending, 2):
    add_row(ws2, i, row_data)

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 15

# ═══════════════════════════════════════════════════════════
# SHEET 3: AUTOMATION STATUS
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Automation Status")
cols3 = 6
headers3 = ["Platform", "Account", "Product", "Automation Script", "Scheduled Task", "Status"]
style_header(ws3, 1, cols3)
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)

auto = [
    ["Reddit", "Khavel_dev", "DevAI", "reddit_post_comment.py", "reddit-warmup-comments (daily 10:26 AM)", "RUNNING"],
    ["Reddit", "StatLineNerd", "NbaPropLab", "TBD (clone reddit_post_comment.py)", "TBD", "NOT BUILT"],
    ["Reddit", "xGNerd", "FutPicks", "TBD (clone reddit_post_comment.py)", "TBD", "NOT BUILT"],
    ["Twitter/X", "@DevAISemanal", "DevAI", "TBD (Playwright)", "TBD", "NOT BUILT"],
    ["Twitter/X", "@StatLineNerd", "NbaPropLab", "TBD (Playwright)", "TBD", "NOT BUILT"],
    ["Twitter/X", "@FutProbLab", "FutPicks", "TBD (Playwright)", "TBD", "NOT BUILT"],
    ["Dev.to", "Khavel", "DevAI", "TBD (REST API cross-post)", "TBD", "NOT BUILT"],
    ["Hashnode", "???", "DevAI", "TBD (GraphQL API cross-post)", "TBD", "NOT BUILT"],
    ["Telegram", "NbaPropLab bot", "NbaPropLab", "NBAv3.Telegram service", "Quartz jobs", "RUNNING"],
    ["Telegram", "FootballLab bot", "FutPicks", "FootballLab.Telegram service", "Quartz jobs", "RUNNING"],
]
for i, row_data in enumerate(auto, 2):
    add_row(ws3, i, row_data)

for i, w in enumerate([14, 22, 14, 42, 48, 18], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── Security warning on Sheet 1 ──
r_warn = ws.max_row + 2
ws.merge_cells(start_row=r_warn, start_column=1, end_row=r_warn, end_column=cols)
cell = ws.cell(row=r_warn, column=1,
    value="WARNING: THIS FILE CONTAINS PASSWORDS AND API KEYS. DO NOT COMMIT TO GIT OR SHARE.")
cell.font = Font(bold=True, color="FF0000", size=12)
cell.fill = warn_fill

# ── Freeze panes ──
ws.freeze_panes = "A2"
ws2.freeze_panes = "A2"
ws3.freeze_panes = "A2"

# Save
outpath = r"C:\Users\ceja_\Desktop\Desarrollos\Spam\devai-newsletter\seo\promotion-accounts-master.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
print(f"Sheets: {wb.sheetnames}")
print(f"Total rows in Accounts Master: {ws.max_row}")
